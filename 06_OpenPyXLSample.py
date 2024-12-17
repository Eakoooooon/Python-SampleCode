"""
openpyxlサンプルコード

pip install openpyxl
"""
import openpyxl

# Excelファイルの読み込み
wb=openpyxl.load_workbook("Book1.xlsx")
# ファイルのシートの1番目を読み込み
ws=wb.worksheets[0]

# 列と行の端っこを取得
max_row=ws.max_row

# データの取得
for row in range(1,max_row,1):
    print(f"レッスン番号：{ws.cell(row=row, column=1).value}　内容：{ws.cell(row=row, column=4).value}")